"""Harvest bugetele TUTUROR UAT-urilor (comune/orase/municipii/sectoare/judete) de la DPFBL/MDLPA.

Sursa: http://www.dpfbl.mdrap.ro/sit_ven_si_chelt_uat.html (cert TLS invalid -> verify=False).
Pagina listeaza ~27 fisiere XLSX/XLS, cate unul pe an (1999-2025), structura "Anexa 24"
("SITUATIA PRIVIND VENITURILE SI CHELTUIELILE pe unitati administrativ-teritoriale").

IA HREF-urile direct din pagina (numele fisierelor sunt inconsistente -> NU construim URL-uri).
Descarca cel mai recent an (2025) si, daca merge, 2024.

Parsare BLOCK-AWARE: layout-ul NU e un tabel flat. Fiecare fisier are 6 sheet-uri:
  - Sheet1/Sheet2 = VENITURI (Sheet2 = detaliere pe consiliu/UAT)
  - Sheet3/Sheet4 = CHELTUIELI pe titluri (Sheet4 = detaliere)
  - Sheet5/Sheet6 = CHELTUIELI pe parti/capitole
Fiecare sheet contine blocuri repetate "Judetul: X" cu propriul header, urmate de:
  TOTAL JUDET / A. Consiliu judetean / B. Total municipii / <UAT-uri> / C. Total orase /
  <UAT-uri> / D. Total comune / <UAT-uri>.
Randurile UAT individuale au forma "<CIF> - MUNICIPIUL/ORAS/COMUNA/SECTORUL <NUME>".
Coloana B (index 1) = total (VENITURI TOTALE in Sheet2, CHELTUIELI TOTALE in Sheet4).

Pentru un layout robust pe ani vechi (.xls cu coloane usor diferite), detectam dinamic
sheet-ul de venituri si cel de cheltuieli dupa textul din header ("VENITURI TOTALE" /
"CHELTUIELI TOTALE") in loc de a presupune indici ficsi de sheet.

Join venituri<->cheltuieli pe CIF (acelasi set de UAT-uri in ambele sheet-uri).

Output: data/v1/bugete/uat.json -> [{uat, judet, cif, tip, an, venituri_lei, cheltuieli_lei}].
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
from typing import Any

import requests
import urllib3

urllib3.disable_warnings()

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

PAGE_URL = "http://www.dpfbl.mdrap.ro/sit_ven_si_chelt_uat.html"
RAW_DIR = os.path.join(ROOT, "data/raw/bugete_uat")
OUT_DIR = os.path.join(ROOT, "data/v1/bugete")
OUT_FILE = os.path.join(OUT_DIR, "uat.json")

# Cati ani sa luam (cei mai recenti). 2 = 2025 + 2024.
N_YEARS = 2

# Rand UAT: "4562923 - MUNICIPIUL ALBA IULIA" / "16397960 - COMUNA CUT" / sectoare Bucuresti.
UAT_ROW = re.compile(r"^(\d{4,})\s*-\s*(.+)$")

# Clasificare tip dupa prefixul numelui.
TIP_PATTERNS = [
    ("sector", re.compile(r"\bsector(ul)?\b", re.I)),
    ("municipiu", re.compile(r"\bmunicipiu", re.I)),
    ("oras", re.compile(r"\bora[sș]", re.I)),
    ("comuna", re.compile(r"\bcomun[ăa]", re.I)),
]


def fetch_page_hrefs() -> list[tuple[int, str]]:
    """Intoarce [(an, url)] pentru fisierele de buget, din HREF-urile reale ale paginii."""
    r = requests.get(PAGE_URL, verify=False, timeout=60)
    r.raise_for_status()
    html = r.text
    hrefs = re.findall(r"href=[\"']([^\"']+)[\"']", html, re.I)
    out: list[tuple[int, str]] = []
    seen: set[str] = set()
    for h in hrefs:
        if not re.search(r"\.xlsx?$", h, re.I):
            continue
        # Extrage anul (4 cifre 19xx/20xx) din numele fisierului.
        m = re.search(r"(19|20)\d{2}", h)
        if not m:
            continue
        an = int(m.group(0))
        url = h if h.lower().startswith("http") else _resolve(h)
        if url in seen:
            continue
        seen.add(url)
        out.append((an, url))
    out.sort(key=lambda t: t[0], reverse=True)
    return out


def _resolve(href: str) -> str:
    from urllib.parse import urljoin

    return urljoin(PAGE_URL, href)


def download(url: str) -> bytes:
    fname = url.rsplit("/", 1)[-1]
    local = os.path.join(RAW_DIR, fname)
    if os.path.exists(local) and os.path.getsize(local) > 10000:
        with open(local, "rb") as fh:
            return fh.read()
    r = requests.get(url, verify=False, timeout=300)
    r.raise_for_status()
    os.makedirs(RAW_DIR, exist_ok=True)
    with open(local, "wb") as fh:
        fh.write(r.content)
    return r.content


def _num(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(" ", "").replace(",", ".")
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
    return None


def _classify(name: str) -> str:
    for tip, pat in TIP_PATTERNS:
        if pat.search(name):
            return tip
    return "necunoscut"


def _iter_xlsx_rows(content: bytes):
    """Iterator (sheet_name, rows) pentru .xlsx via openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for ws in wb.worksheets:  # doar worksheet-uri reale (sare peste Chartsheet etc.)
        if not hasattr(ws, "iter_rows"):
            continue
        yield ws.title, [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()


def _iter_xls_rows(content: bytes):
    """Iterator (sheet_name, rows) pentru .xls vechi via xlrd."""
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    for sh in book.sheets():
        rows = [tuple(sh.row_values(i)) for i in range(sh.nrows)]
        yield sh.name, rows


def _header_kind(rows: list[tuple]) -> str | None:
    """Detecteaza daca un sheet e de VENITURI sau CHELTUIELI dupa textul din primele randuri."""
    blob = " ".join(
        str(c) for r in rows[:14] for c in r[:3] if isinstance(c, str)
    ).upper()
    if "VENITURI TOTALE" in blob:
        return "venituri"
    if "CHELTUIELI TOTALE" in blob:
        return "cheltuieli"
    return None


def _parse_sheet(rows: list[tuple]) -> dict[str, dict]:
    """Parseaza un sheet block-aware -> {cif: {uat, judet, tip, total}}.

    `total` = coloana B (index 1), totalul (venituri sau cheltuieli) al UAT-ului.
    """
    out: dict[str, dict] = {}
    judet = None
    for r in rows:
        # Header de bloc judet.
        for c in r[:3]:
            if isinstance(c, str) and c.strip().startswith("Judetul:"):
                judet = c.split(":", 1)[1].strip()
                break
        a = r[0] if r else None
        if not isinstance(a, str):
            continue
        m = UAT_ROW.match(a.strip())
        if not m:
            continue
        cif, name = m.group(1), re.sub(r"\s+", " ", m.group(2).strip())
        total = _num(r[1]) if len(r) > 1 else None
        if total is None:
            continue
        out[cif] = {
            "cif": cif,
            "uat": name,
            "judet": judet,
            "tip": _classify(name),
            "total": total,
        }
    return out


# ---------------------------------------------------------------------------
# Format B: fisier cu cate un sheet pe judet (folosit pe 2024 si ani similari).
# Sheet-uri numerotate "1".."42" + sheet-uri agregate ("total TARA" etc.).
# In sheet-ul de judet: header "Judeţul: X"; randuri sectiune A/B/C/D; randuri UAT
# cu col0=nr.crt, col1=Cod SIRUTA, col2=NUME, col3=VENITURI TOTALE, col24=CHELTUIELI TOTALE.
# ---------------------------------------------------------------------------
SECTION_RE = re.compile(r"^[ABCD]\.\s", re.I)
SECTION_TIP = {"A": "judet_cj", "B": "municipiu", "C": "oras", "D": "comuna"}


def _looks_like_format_b(sheet_names: list[str]) -> bool:
    nums = sum(1 for s in sheet_names if str(s).strip().isdigit())
    return nums >= 20  # ~42 sheet-uri numerotate, cate unul pe judet


def _parse_county_sheet_b(rows: list[tuple]) -> list[dict]:
    """Parseaza un sheet de judet din formatul B -> list de dict-uri UAT."""
    judet = None
    for r in rows[:8]:
        for c in r[:4]:
            if not isinstance(c, str):
                continue
            if re.match(r"\s*Jude[tţțṭ]ul\b", c, re.I):
                # "Judeţul: ALBA", "Judeţul BACAU", "Judeţul : ILFOV" (cu/fara doua puncte).
                raw = re.sub(r"^\s*Jude[tţțṭ]ul\s*:?\s*", "", c, flags=re.I)
                judet = raw.strip().title()
                break
            if re.match(r"\s*MUNICIPIUL\s+BUCURE[SȘ]TI", c, re.I):
                judet = "Bucuresti"
                break
        if judet:
            break

    # Localizeaza coloanele VENITURI TOTALE / CHELTUIELI TOTALE dupa codurile randului 00.01.02 / 50.02.
    col_ven, col_chelt = 3, 24  # default-uri observate
    for r in rows[:14]:
        for ci, c in enumerate(r):
            s = str(c).strip() if c is not None else ""
            if s == "00.01.02":
                col_ven = ci
            elif s == "50.02" and ci < 30:  # primul 50.02 = cheltuieli totale
                col_chelt = ci

    out: list[dict] = []
    cur_tip = None
    for r in rows:
        c0 = r[0] if r else None
        if isinstance(c0, str) and SECTION_RE.match(c0.strip()):
            cur_tip = SECTION_TIP.get(c0.strip()[0].upper())
            continue
        # Rand UAT: col1 = SIRUTA (numeric), col2 = nume (string).
        siruta = r[1] if len(r) > 1 else None
        name = r[2] if len(r) > 2 else None
        if not isinstance(name, str) or not name.strip():
            continue
        sir = _num(siruta)
        if sir is None:
            continue
        ven = _num(r[col_ven]) if len(r) > col_ven else None
        chelt = _num(r[col_chelt]) if len(r) > col_chelt else None
        nume = re.sub(r"\s+", " ", name.strip())
        out.append(
            {
                "uat": nume,
                "judet": judet,
                "cif": f"siruta:{int(sir)}",
                "tip": cur_tip or _classify(nume),
                "venituri": ven,
                "cheltuieli": chelt,
            }
        )
    return out


def _parse_format_b(iterator) -> list[dict]:
    records: list[dict] = []
    for sn, rows in iterator:
        if not str(sn).strip().isdigit():
            continue  # sare peste agregate ("total TARA", etc.)
        records.extend(_parse_county_sheet_b(rows))
    return records


def parse_file(content: bytes, an: int, is_xls: bool) -> list[dict]:
    """Parseaza un fisier; auto-detecteaza formatul (A = blocuri in 6 sheet-uri, B = sheet/judet)."""
    # Determina numele de sheet-uri pentru a alege formatul.
    if is_xls:
        import xlrd

        book = xlrd.open_workbook(file_contents=content)
        sheet_names = book.sheet_names()
    else:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

    iterator = _iter_xls_rows(content) if is_xls else _iter_xlsx_rows(content)

    if _looks_like_format_b(sheet_names):
        b = _parse_format_b(iterator)
        records: list[dict] = []
        for d in b:
            records.append(
                {
                    "uat": d["uat"],
                    "judet": d["judet"],
                    "cif": d["cif"],
                    "tip": d["tip"],
                    "an": an,
                    "venituri_lei": round(d["venituri"], 2) if d["venituri"] is not None else None,
                    "cheltuieli_lei": round(d["cheltuieli"], 2)
                    if d["cheltuieli"] is not None
                    else None,
                }
            )
        return records

    # Format A: VENITURI in Sheet2, CHELTUIELI in Sheet4, join pe CIF.
    venituri: dict[str, dict] = {}
    cheltuieli: dict[str, dict] = {}
    for sn, rows in iterator:
        kind = _header_kind(rows)
        if kind is None:
            continue
        parsed = _parse_sheet(rows)
        if not parsed:
            continue
        if kind == "venituri" and len(parsed) > len(venituri):
            venituri = parsed
        elif kind == "cheltuieli" and len(parsed) > len(cheltuieli):
            cheltuieli = parsed

    records = []
    all_cifs = set(venituri) | set(cheltuieli)
    for cif in sorted(all_cifs):
        v = venituri.get(cif)
        c = cheltuieli.get(cif)
        base = v or c
        records.append(
            {
                "uat": base["uat"],
                "judet": base["judet"],
                "cif": cif,
                "tip": base["tip"],
                "an": an,
                "venituri_lei": round(v["total"], 2) if v else None,
                "cheltuieli_lei": round(c["total"], 2) if c else None,
            }
        )
    return records


def main() -> dict:
    print(f"GET {PAGE_URL}", flush=True)
    files = fetch_page_hrefs()
    print(f"  {len(files)} fisiere de buget gasite in pagina", flush=True)
    targets = files[:N_YEARS]
    print(f"  tinte: {[(a, u.rsplit('/',1)[-1]) for a,u in targets]}", flush=True)

    all_records: list[dict] = []
    per_year: dict[int, int] = {}
    for an, url in targets:
        print(f"\n== AN {an} == {url}", flush=True)
        try:
            content = download(url)
            print(f"  descarcat {len(content)} bytes", flush=True)
            is_xls = url.lower().endswith(".xls")
            recs = parse_file(content, an, is_xls)
            print(f"  parsate {len(recs)} UAT-uri", flush=True)
            all_records.extend(recs)
            per_year[an] = len(recs)
        except Exception as e:
            print(f"  EROARE an {an}: {e!r}", flush=True)

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT_FILE, "w", encoding="utf-8") as fh:
        json.dump(all_records, fh, ensure_ascii=False, indent=2)
    print(f"\nSCRIS {OUT_FILE}: {len(all_records)} inregistrari", flush=True)
    return {"total": len(all_records), "per_year": per_year, "out": OUT_FILE}


if __name__ == "__main__":
    res = main()
    print(json.dumps(res, ensure_ascii=False), flush=True)
