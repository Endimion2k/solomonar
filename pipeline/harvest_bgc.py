"""Descarcă + parsează BUGETUL GENERAL CONSOLIDAT (BGC) lunar de la Ministerul Finanțelor.

Sursa: URL predictibil, binare /static/ fără captcha:
  https://mfinante.gov.ro/static/10/Mfp/buletin/executii/bgc{DDMMYYYY}.xlsx
unde {DDMMYYYY} = ultima zi a lunii (ex. bgc30042026.xlsx = execuție cumulată ian-apr 2026).
Disponibile lunar 2024-2026 ca .xlsx; 2023 e .xls (404 pe .xlsx) — sărit aici.

Parsare openpyxl: antet pe mai multe rânduri (12-18), coloane neetichetate. Layout fix:
  - col B (idx 2)  = denumire indicator
  - col Q (idx 16) = "Total buget general consolidat" (brut, milioane lei)
  - col R (idx 17) = "Operatiuni financiare" (se scad)
  - col S (idx 18) = "Buget general consolidat - Sume" (NET, după op. financiare) ← headline
  - col S la rândul ~11 conține valoarea PIB anual; col S de la rândul 13 = "% din PIB"
Mapăm manual rândurile cheie după textul etichetei (robustly, nu după nr. rând):
  VENITURI TOTALE, CHELTUIELI TOTALE, EXCEDENT(+)/DEFICIT(-) + sub-clasificație.

Convenția cifrelor: BGC raportează cumulat de la 1 ian; deficitul "Sume" (col S) e cel
raportat oficial (cash). venituri/cheltuieli/sold în câmpurile produse = col S (NET).

Output: data/v1/bugete/bgc.json
  [{perioada, an, luna, venituri_mil_lei, cheltuieli_mil_lei, sold, pib_mil_lei, sold_pct_pib, detalii, source_url}]
"""

from __future__ import annotations

import io
import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "packages", "solomonar_core"))
from solomonar_core.http import Client  # noqa: E402

V = os.path.join(ROOT, "data", "v1", "bugete")
BASE = "https://mfinante.gov.ro/static/10/Mfp/buletin/executii/bgc{}.xlsx"

# Indici coloane (1-based) în layout-ul fix al fișierelor BGC.
COL_LABEL = 2      # B  – denumire indicator
COL_TOTAL = 16     # Q  – Total buget general consolidat (brut)
COL_OPFIN = 17     # R  – Operatiuni financiare (se scad)
COL_NET = 18       # S  – Sume (NET, headline)
COL_PCT = 19       # T  – % din PIB

LUNI = {
    1: "ianuarie", 2: "februarie", 3: "martie", 4: "aprilie", 5: "mai", 6: "iunie",
    7: "iulie", 8: "august", 9: "septembrie", 10: "octombrie", 11: "noiembrie", 12: "decembrie",
}

# (DD, MM, YYYY) pentru ultima zi a fiecărei luni — lunar 2024..2026.
_LAST_DAY = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}


def _leap(y: int) -> bool:
    return y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)


def _candidates() -> list[tuple[int, int, str]]:
    """(an, luna, DDMMYYYY) pentru toate lunile 2024-01 .. 2026-12 (cu zile reale)."""
    out = []
    for y in range(2024, 2027):
        for m in range(1, 13):
            dd = 29 if (m == 2 and _leap(y)) else _LAST_DAY[m]
            out.append((y, m, f"{dd:02d}{m:02d}{y}"))
    return out


def _norm(s) -> str:
    return str(s).replace("\n", " ").strip() if s is not None else ""


def _num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    return None


def _row_label(ws, target_substrings) -> list[int]:
    """Caută indicii rândurilor a căror etichetă (col B) conține un substring (case-insens)."""
    hits = []
    for ri in range(1, ws.max_row + 1):
        lab = _norm(ws.cell(row=ri, column=COL_LABEL).value).upper()
        if lab and any(sub in lab for sub in target_substrings):
            hits.append(ri)
    return hits


def _cell(ws, ri, ci):
    return _num(ws.cell(row=ri, column=ci).value)


def _find_pib(ws):
    """PIB anual: căutat în primele 15 rânduri — celulă text 'PIB <an>' urmată de valoare numerică."""
    for ri in range(1, 16):
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, str) and "PIB" in v.upper():
                nxt = ws.cell(row=ri, column=ci + 1).value
                if isinstance(nxt, (int, float)):
                    return float(nxt)
    return None


def parse_bgc(content: bytes, an: int, luna: int, url: str) -> dict | None:
    """Parsează un fișier xlsx BGC → dict cu indicatorii headline + detalii pe clasificație."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]

    def first(subs):
        rows = _row_label(ws, subs)
        return rows[0] if rows else None

    r_ven = first(["VENITURI TOTALE"])
    r_chl = first(["CHELTUIELI TOTALE"])
    r_sld = first(["EXCEDENT", "DEFICIT"])
    if r_ven is None or r_chl is None or r_sld is None:
        wb.close()
        return None

    venituri = _cell(ws, r_ven, COL_NET)
    cheltuieli = _cell(ws, r_chl, COL_NET)
    sold = _cell(ws, r_sld, COL_NET)
    pib = _find_pib(ws)

    # Detalii: indicatorii principali de clasificație (venit + cheltuieli), col NET + % PIB.
    detalii_labels = [
        "Venituri curente", "Venituri fiscale",
        "Impozitul pe profit", "TVA", "Accize",
        "Contributii de asigurari", "Venituri nefiscale",
        "Sume primite de la UE",
        "Cheltuieli curente", "Cheltuieli de personal", "Bunuri si servicii",
        "Dobanzi", "Asistenta sociala", "Subventii",
        "Cheltuieli de capital", "Active nefinanciare",
    ]
    detalii = []
    used = set()
    for lab_want in detalii_labels:
        for ri in _row_label(ws, [lab_want.upper()]):
            if ri in used:
                continue
            used.add(ri)
            val = _cell(ws, ri, COL_NET)
            if val is None:
                continue
            detalii.append({
                "indicator": _norm(ws.cell(row=ri, column=COL_LABEL).value),
                "mil_lei": val,
                "pct_pib": _cell(ws, ri, COL_PCT),
            })
            break

    wb.close()
    perioada = f"01.01-{_LAST_DAY[luna] if not (luna == 2 and _leap(an)) else 29:02d}.{luna:02d}.{an}"
    return {
        "perioada": perioada,
        "an": an,
        "luna": luna,
        "luna_nume": LUNI[luna],
        "venituri_mil_lei": venituri,
        "cheltuieli_mil_lei": cheltuieli,
        "sold": sold,
        "pib_mil_lei": pib,
        "sold_pct_pib": _cell(ws, r_sld, COL_PCT),
        "detalii": detalii,
        "source_url": url,
    }


def main() -> dict:
    os.makedirs(V, exist_ok=True)
    client = Client(throttle_seconds=0.3, timeout=60)

    cands = _candidates()

    def _fetch(c):
        an, luna, dd = c
        url = BASE.format(dd)
        try:
            r = client.get(url)
            if r.status_code == 200 and r.content[:4] == b"PK\x03\x04":
                return c, url, r.content
        except Exception:
            pass
        return c, url, None

    downloaded = []
    with ThreadPoolExecutor(max_workers=8) as ex:
        for c, url, content in ex.map(_fetch, cands):
            if content is not None:
                downloaded.append((c, url, content))
    print(f"fișiere xlsx găsite: {len(downloaded)} / {len(cands)} candidați", flush=True)

    records = []
    for (an, luna, dd), url, content in downloaded:
        try:
            rec = parse_bgc(content, an, luna, url)
            if rec is not None:
                records.append(rec)
            else:
                print(f"  PARSE-MISS {dd}: rânduri headline negăsite", flush=True)
        except Exception as e:
            print(f"  PARSE-ERR {dd}: {e!r}", flush=True)

    records.sort(key=lambda r: (r["an"], r["luna"]))

    out = os.path.join(V, "bgc.json")
    json.dump(records, open(out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"scris {len(records)} înregistrări -> {out}", flush=True)

    # Eșantion verificare
    for r in records[-3:]:
        print(
            f"  {r['perioada']}: V={r['venituri_mil_lei']:.0f} C={r['cheltuieli_mil_lei']:.0f} "
            f"sold={r['sold']:.0f} ({r['sold_pct_pib']}% PIB) PIB={r['pib_mil_lei']}",
            flush=True,
        )
    return {"records": len(records), "out": out}


if __name__ == "__main__":
    main()
